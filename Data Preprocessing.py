# -*- coding: utf-8 -*-
"""
Created on Tue Jul  2 12:22:22 2024

@author: Shiyun Liu
"""
import numpy as np
import yfinance as yf
import pandas as pd
import openpyxl
# %% fetch stock data

# Create the Ticker object
aapl = yf.Ticker("AAPL")
# Get historical market data
# hist = aapl.history(start="2022-10-01", end="2024-04-02")
hist = aapl.history(start="2014-07-02", end="2024-07-02")
income_stmt = aapl.quarterly_income_stmt.T
balance_sheet = aapl.quarterly_balance_sheet.T
cashflow = aapl.quarterly_cashflow.T

# %%
# Remove timezone information
hist.index = hist.index.tz_localize(None)

# Save the data to an Excel file
hist.to_excel("AAPL_stockprices.xlsx")


with pd.ExcelWriter("AAPL_financials.xlsx") as writer:
    income_stmt.to_excel(writer, sheet_name="quarterly_income_stmt")
    balance_sheet.to_excel(writer, sheet_name="quarterly_balance_sheet")
    cashflow.to_excel(writer, sheet_name="quarterly_cashflow")


# %% Extract 'Date' and 'High', 'Dividends', 'Stock Splits' columns from "AAPL_stockprices.xlsx" to "AAPL_input.xlsx"

file_path = 'AAPL_financials.xlsx'

# Load the workbook
wb = openpyxl.load_workbook(file_path)

# Iterate through all sheets
for sheet_name in wb.sheetnames:
    # Select the sheet
    sheet = wb[sheet_name]

    # Rename the existing first cell to 'Date'
    sheet.cell(row=1, column=1, value='Date')

    # Save the workbook
    wb.save(file_path)

    print(f'Heading "Date" added successfully to {sheet_name}.')

print('All sheets updated successfully.')

# %% Replace empty cells in all sheets in "AAPL_financials.xlsx" with the numerical value 0
# Load the Excel file
# Load the Excel file
file_path = "AAPL_financials.xlsx"

# Load the Excel file into a dictionary of DataFrames, one for each sheet
xls = pd.read_excel(file_path, sheet_name=None)

# Iterate through each sheet and replace empty cells with 0
for sheet_name, df in xls.items():
    df.fillna(0, inplace=True)

# Save the modified Excel file
with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
    for sheet_name, df in xls.items():
        # Write each DataFrame to Excel
        df.to_excel(writer, sheet_name=sheet_name, index=False)

print("Empty cells replaced with 0 and saved to AAPL_financials.xlsx")


# %% Append Data to "AAPL_input.xlsx"
# Step 1: Read the Excel files into DataFrames
input_file = "AAPL_stockprices.xlsx"
financials_file = "AAPL_financials.xlsx"

# Read each sheet into separate DataFrames
xls = pd.ExcelFile(financials_file)
df_income_stmt = pd.read_excel(xls, 'quarterly_income_stmt')
df_balance_sheet = pd.read_excel(xls, 'quarterly_balance_sheet')
df_cashflow = pd.read_excel(xls, 'quarterly_cashflow')

# Read the input data
df_input = pd.read_excel(input_file)

# Step 2: Define a function to check if a date is within a quarter


def is_within_quarter(date, quarter_date):
    if quarter_date.month == 3 and quarter_date.day == 31:
        return date.month in [1, 2, 3]
    elif quarter_date.month == 6 and quarter_date.day == 30:
        return date.month in [4, 5, 6]
    elif quarter_date.month == 9 and quarter_date.day == 30:
        return date.month in [7, 8, 9]
    elif quarter_date.month == 12 and quarter_date.day == 31:
        return date.month in [10, 11, 12]
    else:
        return False


# Step 3: Merge all financial data based on quarter match
appended_data = []

for index, row in df_input.iterrows():
    input_date = row['Date']

    for idx, fin_row in df_income_stmt.iterrows():
        financial_date = fin_row['Date']
        if is_within_quarter(input_date, financial_date):
            merged_row = {**row, **fin_row.drop('Date')}
            break

    for idx, fin_row in df_balance_sheet.iterrows():
        financial_date = fin_row['Date']
        if is_within_quarter(input_date, financial_date):
            merged_row.update(fin_row.drop('Date'))
            break

    for idx, fin_row in df_cashflow.iterrows():
        financial_date = fin_row['Date']
        if is_within_quarter(input_date, financial_date):
            merged_row.update(fin_row.drop('Date'))
            break

    appended_data.append(merged_row)

# Step 4: Convert the appended data to DataFrame and write to Excel
df_combined = pd.DataFrame(appended_data)

df_combined.iloc[:-1].to_excel("AAPL_input.xlsx", index=False)

# %% Extract 'Close' columns from "AAPL_stockprices.xlsx" to "AAPL_output.xlsx"

# Step 1: Read the data from AAPL_stockprices.xlsx
stockprices_file = "AAPL_stockprices.xlsx"
stockprices_df = pd.read_excel(stockprices_file)

# Step 2: Extract 'Close' column
extracted_data = stockprices_df[['Close']]

# Step 3: Write the extracted data to AAPL_output.xlsx
output_file = "AAPL_output.xlsx"
extracted_data.iloc[1:].to_excel(output_file, index=False)

print(f"Extracted data saved to {output_file}")


# %% binary classification
# Load the Excel file
file_path = 'AAPL_stockprices.xlsx'
df = pd.read_excel(file_path)

# Extract the 'Close' column
close_column = df['Close']

# Calculate differences and update the column
close_column_diff = close_column.diff()
# Assign differences starting from the second row
close_column.iloc[1:] = close_column_diff.iloc[1:]

# Update the original DataFrame with the modified 'Close' column
df['Close'] = close_column
# Modify 'Close' column based on conditions
df['Close'] = (df['Close'] > 0).astype(int)
# Save the updated DataFrame to Excel
updated_file_path = 'AAPL_output.xlsx'
df['Close'].iloc[1:].to_excel(updated_file_path, index=False)
